import openpyxl

book = openpyxl.load_workbook("C:\\Users\\kamal.kumar\\Desktop\\PythonDemo.xlsx")
sheet = book.active
Dict = {}
cell = sheet.cell(row=1, column=2)
print(cell.value)
sheet.cell(row=2, column=2).value = "Kamal"
print(sheet.cell(row=2, column=2).value)
print(sheet.max_row)
print(sheet.max_column)
print(sheet["A5"].value)

